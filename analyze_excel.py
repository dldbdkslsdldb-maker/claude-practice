import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(r'c:\Users\유창현\Desktop\claude-practice\챕터3숙제_데이터분석.xlsx')
ws = wb['월별판매데이터']

headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
rows = []
for r in range(2, ws.max_row+1):
    row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column+1)}
    rows.append(row)

print(f'총 데이터 건수: {len(rows)}건')
print(f'컬럼 수: {len(headers)}개')
print(f'컬럼 목록: {headers}')
print()

# 데이터 유형 분석
print('=== 컬럼별 데이터 유형 ===')
for col in headers:
    vals = [r[col] for r in rows if r[col] is not None]
    types = set(type(v).__name__ for v in vals)
    print(f'  {col}: {types} (비어있음: {len(rows)-len(vals)}개)')

print()

# 날짜 범위
dates = [r['주문일자'] for r in rows if r['주문일자']]
print('=== 날짜 범위 ===')
print(f'  시작: {min(dates).strftime("%Y-%m-%d")}')
print(f'  종료: {max(dates).strftime("%Y-%m-%d")}')
months = Counter(d.month for d in dates)
print(f'  월별 건수: {dict(sorted(months.items()))}')

print()

# 지역별
regions = Counter(r['지역'] for r in rows)
print('=== 지역별 건수 ===')
for k, v in regions.most_common():
    print(f'  {k}: {v}건')

print()

# 카테고리별
cats = Counter(r['카테고리'] for r in rows)
print('=== 카테고리별 건수 ===')
for k, v in cats.most_common():
    print(f'  {k}: {v}건')

print()

# 판매채널별
channels = Counter(r['판매채널'] for r in rows)
print('=== 판매채널별 건수 ===')
for k, v in channels.most_common():
    print(f'  {k}: {v}건')

print()

# 수량 통계
qtys = [r['수량'] for r in rows if isinstance(r['수량'], (int, float))]
print('=== 수량 통계 ===')
print(f'  최소: {min(qtys)}, 최대: {max(qtys)}, 평균: {sum(qtys)/len(qtys):.2f}')

# 단가 통계
prices = [r['단가'] for r in rows if isinstance(r['단가'], (int, float))]
print('=== 단가 통계 ===')
print(f'  최소: {min(prices):,}원')
print(f'  최대: {max(prices):,}원')
print(f'  평균: {sum(prices)/len(prices):,.0f}원')

# 상품명 목록
products = Counter(r['상품명'] for r in rows)
print()
print(f'=== 전체 고유 상품 수: {len(products)}개 ===')
print('  상위 5개 상품:')
for k, v in products.most_common(5):
    print(f'    {k}: {v}건')

# 매출 컬럼 형식 확인
print()
print('=== 매출 컬럼 샘플 ===')
for i, r in enumerate(rows[:5]):
    print(f'  행{i+2}: {r["매출"]}')

# 월별 카테고리 크로스 분석
print()
print('=== 월별 카테고리 건수 ===')
month_cat = {}
for r in rows:
    m = r['주문일자'].month if r['주문일자'] else None
    cat = r['카테고리']
    if m and cat:
        key = (m, cat)
        month_cat[key] = month_cat.get(key, 0) + 1

cat_list = list(set(r['카테고리'] for r in rows if r['카테고리']))
month_list = sorted(set(r['주문일자'].month for r in rows if r['주문일자']))
header_row = '월  | ' + ' | '.join(f'{c[:4]}' for c in cat_list)
print(f'  {header_row}')
for m in month_list:
    vals = [str(month_cat.get((m, c), 0)).rjust(4) for c in cat_list]
    print(f'  {m}월 | ' + ' | '.join(vals))
