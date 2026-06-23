import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from collections import defaultdict
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import matplotlib.ticker as mticker

# ── 한글 폰트 자동 탐지 ───────────────────────────────────
available = [f.name for f in fm.fontManager.ttflist]
for name in ['Malgun Gothic', 'NanumGothic', 'AppleGothic']:
    if name in available:
        plt.rcParams['font.family'] = name
        print(f'[폰트] {name} 적용')
        break
else:
    print('[폰트] 한글 폰트 미탐지 — 기본 폰트 사용')
plt.rcParams['axes.unicode_minus'] = False

# ── 1단계: 데이터 로드 ────────────────────────────────────
FILE = r'c:\Users\유창현\Desktop\claude-practice\챕터3숙제_데이터분석.xlsx'
wb = openpyxl.load_workbook(FILE)
ws = wb['월별판매데이터']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

rows = []
for r in range(2, ws.max_row + 1):
    row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
    # ⚠️ 수식 셀 감지 → 수량 × 단가로 직접 계산
    if isinstance(row.get('매출'), str) and row['매출'].startswith('='):
        row['매출액'] = (row['수량'] or 0) * (row['단가'] or 0)
    else:
        row['매출액'] = row.get('매출') or 0
    rows.append(row)

print(f'[구조] {len(rows)}건 × {len(headers)}개 컬럼 로드 완료')
print(f'[수식] 매출 컬럼 수식 감지 → 수량×단가로 재계산 완료')

# ── 2단계: 집계 ───────────────────────────────────────────
MONTHS      = [1, 2, 3, 4]
MONTH_LABEL = ['1월', '2월', '3월', '4월']
CATS        = ['전자기기', '생활가전', '주방용품', '뷰티']
CHANNELS    = ['쿠팡', '자사몰', '네이버스마트스토어', '11번가', '오프라인매장']

# 월별 × 카테고리 매출
month_cat = defaultdict(lambda: defaultdict(int))
for r in rows:
    month_cat[r['주문일자'].month][r['카테고리']] += r['매출액']

month_total = {m: sum(month_cat[m].values()) for m in MONTHS}

# 판매채널별 매출
channel_sales = defaultdict(int)
for r in rows:
    channel_sales[r['판매채널']] += r['매출액']

# 지역별 건수
region_cnt = defaultdict(int)
for r in rows:
    region_cnt[r['지역']] += 1
regions_sorted = sorted(region_cnt.items(), key=lambda x: -x[1])

# ── 색상 팔레트 ───────────────────────────────────────────
CAT_CLR = ['#4C72B0', '#DD8452', '#55A868', '#C44E52']
CH_CLR  = ['#4C72B0', '#55A868', '#C44E52', '#8172B2', '#937860']

OUT = r'c:\Users\유창현\Desktop\claude-practice'

# ══════════════════════════════════════════════════════════
# 차트 1 ── 선 차트: 월별 카테고리별 매출 추이
# ══════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(9, 5))
for i, cat in enumerate(CATS):
    vals = [month_cat[m][cat] / 1e6 for m in MONTHS]
    ax.plot(MONTH_LABEL, vals, marker='o', markersize=7,
            linewidth=2.2, color=CAT_CLR[i], label=cat)
    for j, v in enumerate(vals):
        ax.annotate(f'{v:.1f}M', (MONTH_LABEL[j], v),
                    textcoords='offset points', xytext=(0, 9),
                    ha='center', fontsize=8, color=CAT_CLR[i])

totals = [month_total[m] / 1e6 for m in MONTHS]
ax.plot(MONTH_LABEL, totals, marker='s', markersize=6, linewidth=1.5,
        linestyle='--', color='gray', label='합계', alpha=0.7)
for j, v in enumerate(totals):
    ax.annotate(f'{v:.1f}M', (MONTH_LABEL[j], v),
                textcoords='offset points', xytext=(0, 9),
                ha='center', fontsize=8, color='gray')

ax.set_title('월별 카테고리별 매출 추이', fontsize=14, fontweight='bold', pad=14)
ax.set_ylabel('매출 (백만원)', fontsize=11)
ax.set_xlabel('월', fontsize=11)
ax.set_ylim(0)
ax.legend(loc='upper left', fontsize=9)
ax.spines[['top', 'right']].set_visible(False)
ax.grid(axis='y', linestyle='--', alpha=0.4)
fig.tight_layout()
fig.savefig(f'{OUT}/chart_line.png', dpi=150, bbox_inches='tight')
plt.close()
print('chart_line.png 저장 완료')

# ══════════════════════════════════════════════════════════
# 차트 2 ── 누적 막대 차트: 월별 카테고리별 매출
# ══════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(9, 5))
bottoms = [0.0] * 4
for i, cat in enumerate(CATS):
    vals = [month_cat[m][cat] / 1e6 for m in MONTHS]
    bars = ax.bar(MONTH_LABEL, vals, bottom=bottoms,
                  color=CAT_CLR[i], label=cat,
                  width=0.5, edgecolor='white', linewidth=0.8)
    for j, (b, v) in enumerate(zip(bottoms, vals)):
        if v >= 1.0:
            ax.text(j, b + v / 2, f'{v:.1f}', ha='center', va='center',
                    fontsize=8, color='white', fontweight='bold')
    bottoms = [bottoms[k] + vals[k] for k in range(4)]

ax.set_title('월별 카테고리별 매출 (누적 막대)', fontsize=14, fontweight='bold', pad=14)
ax.set_ylabel('매출 (백만원)', fontsize=11)
ax.set_xlabel('월', fontsize=11)
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:.0f}M'))
ax.legend(loc='upper right', fontsize=9)
ax.spines[['top', 'right']].set_visible(False)
ax.grid(axis='y', linestyle='--', alpha=0.4)
fig.tight_layout()
fig.savefig(f'{OUT}/chart_bar.png', dpi=150, bbox_inches='tight')
plt.close()
print('chart_bar.png 저장 완료')

# ══════════════════════════════════════════════════════════
# 차트 3 ── 원형 차트: 판매채널별 매출 비중
# ══════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(7, 6))
ch_vals = [channel_sales[ch] / 1e6 for ch in CHANNELS]
wedges, texts, autotexts = ax.pie(
    ch_vals,
    labels=CHANNELS,
    colors=CH_CLR,
    explode=[0.04] * len(CHANNELS),
    autopct='%1.1f%%',
    startangle=140,
    pctdistance=0.78,
    textprops={'fontsize': 10},
)
for at in autotexts:
    at.set_fontsize(9)
    at.set_color('white')
    at.set_fontweight('bold')

legend_labels = [f'{ch}  ({v:.1f}M)' for ch, v in zip(CHANNELS, ch_vals)]
ax.legend(wedges, legend_labels, loc='lower center',
          bbox_to_anchor=(0.5, -0.14), ncol=2, fontsize=9)
ax.set_title('판매채널별 매출 비중', fontsize=14, fontweight='bold', pad=14)
fig.tight_layout()
fig.savefig(f'{OUT}/chart_pie.png', dpi=150, bbox_inches='tight')
plt.close()
print('chart_pie.png 저장 완료')

# ══════════════════════════════════════════════════════════
# 차트 4 ── 수평 막대 차트: 지역별 주문 건수
# ══════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(8, 4.5))
reg_names = [r[0] for r in regions_sorted]
reg_vals  = [r[1] for r in regions_sorted]
bar_colors = ['#4C72B0' if i == 0 else '#A8C0DD' for i in range(len(reg_names))]

hbars = ax.barh(reg_names, reg_vals, color=bar_colors,
                edgecolor='white', height=0.55)
for bar, v in zip(hbars, reg_vals):
    ax.text(v + 0.5, bar.get_y() + bar.get_height() / 2,
            f'{v}건', va='center', fontsize=10)

ax.set_title('지역별 주문 건수', fontsize=14, fontweight='bold', pad=14)
ax.set_xlabel('주문 건수', fontsize=11)
ax.invert_yaxis()
ax.set_xlim(0, max(reg_vals) * 1.2)
ax.spines[['top', 'right', 'left']].set_visible(False)
ax.tick_params(left=False)
ax.grid(axis='x', linestyle='--', alpha=0.4)
fig.tight_layout()
fig.savefig(f'{OUT}/chart_hbar.png', dpi=150, bbox_inches='tight')
plt.close()
print('chart_hbar.png 저장 완료')

# ── 참고 수치 출력 ────────────────────────────────────────
print('\n--- 참고 수치 ---')
print('월별 합계(M):', {f'{m}월': round(month_total[m]/1e6, 1) for m in MONTHS})
print('채널별 매출(M):', {ch: round(channel_sales[ch]/1e6, 1) for ch in CHANNELS})
total_all = sum(r['매출액'] for r in rows)
print(f'전체 총매출: {total_all/1e6:.1f}M원')
print(f'평균 단가: {sum(r["단가"] for r in rows)/len(rows):,.0f}원')
